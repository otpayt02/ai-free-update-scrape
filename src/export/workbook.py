"""Pretty local workbook for Sheets import and live dashboard review."""

from __future__ import annotations

from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "0B1020"
SLATE = "121933"
BLUE = "5EA1FF"
CYAN = "8DD6FF"
GREEN = "4ADE80"
AMBER = "FBBF24"
RED = "F87171"
WHITE = "F8FAFC"
GRID = "243055"


def _style_header(row):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=SLATE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=GRID))


def _autosize(ws):
    for column_cells in ws.columns:
        values = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
        width = min(max(values, default=10) + 2, 48)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _write_table(ws, start_row: int, start_col: int, rows):
    for row_index, row in enumerate(rows, start=start_row):
        for col_index, value in enumerate(row, start=start_col):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_dashboard_workbook(
    articles: list[dict],
    shorts_plan: list[dict],
    digest_path: Path,
    output_path: Path,
) -> Path:
    """Build a styled dashboard workbook from the current scrape outputs."""

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_articles = wb.create_sheet("Articles")
    ws_shorts = wb.create_sheet("Shorts Plan")
    ws_trends = wb.create_sheet("Trends")
    ws_notes = wb.create_sheet("How To Use")

    today = date.today().isoformat()
    top_articles = sorted(articles, key=lambda a: a.get("ranking", {}).get("top_score", 0), reverse=True)[:12]
    topic_counts = Counter(topic for article in articles for topic in article.get("topics", []))
    source_counts = Counter(article.get("source", "Unknown") for article in articles)

    ws_dash.sheet_view.showGridLines = False
    ws_dash["A1"] = "AI Free Update Scrape"
    ws_dash["A1"].font = Font(size=24, bold=True, color=WHITE)
    ws_dash["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws_dash["A2"] = f"Generated {today}"
    ws_dash["A2"].font = Font(color=CYAN, italic=True)
    ws_dash["A4"] = "Digest"
    ws_dash["B4"] = str(digest_path.name)
    ws_dash["A5"] = "Articles"
    ws_dash["B5"] = len(articles)
    ws_dash["A6"] = "Shorts rows"
    ws_dash["B6"] = len(shorts_plan)
    ws_dash["A7"] = "Top source"
    ws_dash["B7"] = source_counts.most_common(1)[0][0] if source_counts else "n/a"

    for cell in ("A4", "A5", "A6", "A7"):
        ws_dash[cell].font = Font(bold=True, color=WHITE)
    for cell in ("B4", "B5", "B6", "B7"):
        ws_dash[cell].font = Font(bold=True, color=GREEN)

    ws_dash["D4"] = "Topic mix"
    ws_dash["D4"].font = Font(bold=True, color=WHITE)
    row = 5
    for topic, count in topic_counts.most_common(8):
        ws_dash.cell(row=row, column=4, value=topic)
        ws_dash.cell(row=row, column=5, value=count)
        row += 1

    chart = BarChart()
    chart.title = "Topic Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Topic"
    data = Reference(ws_dash, min_col=5, min_row=5, max_row=max(row - 1, 5))
    cats = Reference(ws_dash, min_col=4, min_row=5, max_row=max(row - 1, 5))
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    ws_dash.add_chart(chart, "G4")

    ws_dash["D16"] = "Top picks"
    ws_dash["D16"].font = Font(bold=True, color=WHITE)
    pick_headers = ["Score", "Title", "Source", "Topics"]
    _write_table(ws_dash, 17, 4, [pick_headers] + [
        [
            article.get("ranking", {}).get("top_score", 0),
            article.get("title", ""),
            article.get("source", ""),
            ", ".join(article.get("topics", [])),
        ]
        for article in top_articles
    ])
    _style_header(ws_dash[17][3:7])

    ws_articles.append([
        "published", "scraped_at", "source", "title", "url", "summary", "topics", "new_tool", "tool_type", "tool_name", "top_score", "reason"
    ])
    for article in articles:
        detection = article.get("detection", {})
        ranking = article.get("ranking", {})
        ws_articles.append([
            article.get("published", ""),
            article.get("scraped_at", ""),
            article.get("source", ""),
            article.get("title", ""),
            article.get("url", ""),
            article.get("summary", ""),
            ", ".join(article.get("topics", [])),
            detection.get("new_tool", False),
            detection.get("type", ""),
            detection.get("tool_name", ""),
            ranking.get("top_score", 0),
            ranking.get("reason", ""),
        ])
    _style_header(ws_articles[1])

    ws_shorts.append(["publish_date", "slot", "angle", "hook", "source_title", "url", "topics", "score", "summary", "free_angle"])
    for row in shorts_plan:
        ws_shorts.append([
            row.get("publish_date", ""),
            row.get("slot", ""),
            row.get("angle", ""),
            row.get("hook", ""),
            row.get("source_title", ""),
            row.get("url", ""),
            row.get("topics", ""),
            row.get("score", 0),
            row.get("summary", ""),
            row.get("free_angle", ""),
        ])
    _style_header(ws_shorts[1])

    ws_trends.append(["source", "count"])
    for source, count in source_counts.most_common():
        ws_trends.append([source, count])
    _style_header(ws_trends[1])

    line = LineChart()
    line.title = "Source Volume"
    line.y_axis.title = "Count"
    line.x_axis.title = "Source"
    data = Reference(ws_trends, min_col=2, min_row=1, max_row=ws_trends.max_row)
    cats = Reference(ws_trends, min_col=1, min_row=2, max_row=ws_trends.max_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.height = 7
    line.width = 12
    ws_trends.add_chart(line, "D2")

    notes = [
        ["How To Use"],
        ["1. Open Dashboard for live visual review."],
        ["2. Import this workbook as a native Google Sheet."],
        ["3. Use Articles to inspect the scrape feed."],
        ["4. Use Shorts Plan to power your production queue."],
        ["5. Refresh the workbook after each scrape run."],
    ]
    _write_table(ws_notes, 1, 1, notes)
    ws_notes["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws_notes["A1"].fill = PatternFill("solid", fgColor=NAVY)

    for ws in [ws_dash, ws_articles, ws_shorts, ws_trends, ws_notes]:
        ws.sheet_view.zoomScale = 90
        ws.freeze_panes = "A2"
        _autosize(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
